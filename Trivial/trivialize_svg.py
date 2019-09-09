#!/usr/bin/env python3

# Author: Angel Berlanas Vicente
# Email: <angel.berlanas@gmail.com>

# This script is licensed under GPL v3 or higher

import openpyxl
import os
import sys

from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw 

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

# Some Values

numTarjetas=30
trivial_path="TrivialISO.xlsx"
# Positions - respecto al ancho - alto
# qbpos = (ancho,alto)
#qbpos = [(127,50),(127,115),(127,175),(127,240),(127,300),(127,370)]
#rbpos = [(380,50),(380,115),(380,175),(380,240),(380,300),(380,370)]
qbpos = [(64,27),(64,60),(64,93),(64,128),(64,163),(64,195)]
rbpos = [(190,27),(190,60),(190,93),(190,128),(190,163),(190,195)]


wb = openpyxl.load_workbook(filename = trivial_path)
ws = wb.active



for evaluacion in ("1","2","3"):
    
    i=1
    
    while (i<=numTarjetas):

        pos=0
        #img_p = Image.open("anverso_carta.png")
        #draw_p = ImageDraw.Draw(img_p)

        #img_r = Image.open("reverso_"+evaluacion+".png")
        #draw_r = ImageDraw.Draw(img_r)
    
        with open('dibujo_preguntas.svg', 'r') as filep :
            filedatap = filep.read()
        
        with open('dibujo_respuesta_'+evaluacion+'.svg', 'r') as filer :
            filedatar = filer.read()

        for hoja in (wb.sheetnames):
            if (str(hoja).find('_'+evaluacion) != -1):
                color = hoja.split("_")[0]
                #print (color)
                ws = wb[hoja]

                qb = str(ws['A'+str(i)].value)
                rb = str(ws['B'+str(i)].value)

                if (len(qb) > 57):
                    print("Hoja :"+hoja+" Fila: "+str(i)+" -- Pregunta * : "+ qb)


                if (len(rb) > 29):
                   print("Hoja :"+hoja+" Fila: "+str(i)+" -- Respuesta * : "+ rb)

                # Pregunta
                # Replace the target string
                filedatap = filedatap.replace(color, qb)


                # Respuesta
                #draw_r.text(rbpos[pos],rb,(0,0,0),font=font)

                    # Replace the target string
                filedatar = filedatar.replace(color, rb)

                # Write the file out again
                #with open('dibujo_test_relleno.svg', 'w') as file:
                #    file.write(filedata)
                
                

        if not os.path.exists("Eva"+evaluacion+"Carta_"+str(i)):
            os.mkdir("Eva"+evaluacion+"Carta_"+str(i))
           
        with open('Eva'+evaluacion+'Carta_'+str(i)+'/'+"Eva"+evaluacion+'Carta_'+str(i)+'_CARA_A.svg', 'w') as file:
            file.write(filedatap)
        with open('Eva'+evaluacion+'Carta_'+str(i)+'/'+"Eva"+evaluacion+'Carta_'+str(i)+'_CARA_B.svg', 'w') as file:
            file.write(filedatar)        
        #img_p.save("Eva"+evaluacion+'Carta_'+str(i)+'/'+"Eva"+evaluacion+'Carta_'+str(i)+'_CARA_A.png', format="PNG", quality=100, optimize=True, progressive=True)
        #img_r.save("Eva"+evaluacion+'Carta_'+str(i)+'/'+"Eva"+evaluacion+'Carta_'+str(i)+'_CARA_B.png', format="PNG", quality=100, optimize=True, progressive=True)

        i=i+1


